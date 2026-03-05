"""Generate Excel export with 5 columns: Name, Email, Phone, Score, Explanation."""
from __future__ import annotations

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.schemas import CandidateResult


_HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_DATA_FONT = Font(name="Calibri", size=10)
_ALT_FILL = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
_THIN = Side(border_style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

HEADERS = ["Name", "Email", "Phone", "Score", "Explanation"]
COL_WIDTHS = [28, 34, 18, 10, 80]


def generate_excel(results: list[CandidateResult]) -> bytes:
    """
    Build an Excel workbook from candidate results sorted by score (descending).
    Returns raw bytes ready to be streamed as a download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate Rankings"

    # ── Header row ─────────────────────────────────────────────────────────────
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    ws.row_dimensions[1].height = 22

    # ── Data rows ──────────────────────────────────────────────────────────────
    sorted_results = sorted(
        [r for r in results if r.stage == "done"],
        key=lambda r: (r.score or 0),
        reverse=True,
    )

    for row_idx, candidate in enumerate(sorted_results, start=2):
        name = candidate.name or "—"
        email = candidate.email or "—"
        phone = candidate.phone or "—"
        score = candidate.score if candidate.score is not None else "—"
        explanation = _merge_explanation(candidate.explanation)

        row_data = [name, email, phone, score, explanation]
        fill = _ALT_FILL if row_idx % 2 == 0 else None

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _DATA_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
                horizontal="center" if col_idx == 4 else "left",
            )
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 60

    # ── Column widths ──────────────────────────────────────────────────────────
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze header ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _merge_explanation(bullets: list[str]) -> str:
    if not bullets:
        return "—"
    return "\n".join(f"• {b.lstrip('•').strip()}" for b in bullets)
